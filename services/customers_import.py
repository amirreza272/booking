"""
services/customers_import.py
───────────────────────────────
وارد کردن دسته‌ای مشتریان از فایل اکسل به دیتابیس + گوگل‌شیت.

فرمت مورد انتظار فایل (بدون نیاز به هدر، ولی اگه هدر هم داشته باشه
خودکار ردش می‌کنه چون شماره‌ی معتبر نداره):
ستون A: نام | ستون B: شماره | ستون C: تاریخ تولد | ستون D: شهر
"""

import re
from openpyxl import load_workbook

from extensions import db
from models import Customer
from services.gsheet_customers import _get_worksheet, upsert_customer


def _normalize_phone(value):
    if value is None:
        return ""

    # اکسل معمولاً شماره رو به‌عنوان عدد ذخیره می‌کنه (مثلاً 912345678.0)
    # و صفر اولش رو هم گم می‌کنه — این‌ها رو درست می‌کنیم.
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    s = re.sub(r"\D", "", str(value).strip())
    if s and not s.startswith("0"):
        s = "0" + s
    return s


def import_customers_from_excel(file_stream):
    """
    file_stream: فایل‌آبجکت آپلودی (مثل request.files['file'])
    خروجی: {"total": ..., "added": ..., "updated": ..., "errors": [...]}
    هرگز نباید کل عملیات رو با یه ردیف خراب متوقف کنه — ردیف بد رد می‌شه
    و توی errors ثبت می‌شه.
    """
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active

    report = {"total": 0, "added": 0, "updated": 0, "errors": []}
    rows = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        name_raw = row[0] if len(row) > 0 else None
        phone_raw = row[1] if len(row) > 1 else None
        birthdate_raw = row[2] if len(row) > 2 else None
        city_raw = row[3] if len(row) > 3 else None

        phone = _normalize_phone(phone_raw)

        # ردیف‌هایی که شماره‌ی معتبر ندارن رد می‌شن (شامل ردیف هدر، اگه باشه)
        if not (phone.startswith("09") and len(phone) == 11 and phone.isdigit()):
            if name_raw or phone_raw:
                report["errors"].append(f"ردیف رد شد (شماره نامعتبر): {row}")
            continue

        name = str(name_raw).strip() if name_raw else "بدون نام"
        birthdate = str(birthdate_raw).strip() if birthdate_raw else ""
        city = str(city_raw).strip() if city_raw else ""

        rows.append({"phone": phone, "name": name, "birthdate": birthdate, "city": city})

    report["total"] = len(rows)

    existing = {c.phone: c for c in Customer.query.all()}
    new_for_sheet = []
    updated_customers = []

    for item in rows:
        if item["phone"] in existing:
            c = existing[item["phone"]]
            c.name = item["name"]
            c.birthdate = item["birthdate"]
            c.city = item["city"]
            report["updated"] += 1
            updated_customers.append(c)
        else:
            c = Customer(phone=item["phone"], name=item["name"], birthdate=item["birthdate"], city=item["city"])
            db.session.add(c)
            existing[item["phone"]] = c
            report["added"] += 1
            new_for_sheet.append(item)

    db.session.commit()

    # ─── گوگل‌شیت ───
    # مشتری‌های *جدید* رو یکجا (یک درخواست API) اضافه می‌کنیم — برای فایل‌های
    # بزرگ خیلی سریع‌تره. اگه این روش دسته‌ای به هر دلیلی خطا داد، به روش
    # تک‌تک (که قبلاً برای افزودن دستی تست و تأیید شده) fallback می‌کنیم
    # تا داده هیچ‌وقت گم نشه.
    if new_for_sheet:
        try:
            worksheet = _get_worksheet()
            values = [[i["phone"], i["name"], i["birthdate"], i["city"]] for i in new_for_sheet]
            worksheet.append_rows(values, value_input_option="USER_ENTERED")
        except Exception as e:
            report["errors"].append(f"افزودن دسته‌ای به گوگل‌شیت شکست خورد ({e}) — در حال تلاش تک‌تک...")
            for i in new_for_sheet:
                try:
                    worksheet = _get_worksheet()
                    worksheet.append_row(
                        [i["phone"], i["name"], i["birthdate"], i["city"]],
                        value_input_option="USER_ENTERED",
                    )
                except Exception as e2:
                    report["errors"].append(f"خطا در افزودن مشتری {i['phone']} به گوگل‌شیت: {e2}")

    # مشتری‌های آپدیت‌شده (که از قبل توی شیت هم بودن) رو تک‌تک sync می‌کنیم —
    # معمولاً توی یه وارد کردن اولیه تعدادشون کمه یا صفره.
    for c in updated_customers:
        upsert_customer(c)

    return report